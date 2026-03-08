import pandas as pd
import os
import glob
from openpyxl import load_workbook

def limpiar_datos():
    archivos = glob.glob('impos*.*')
    if not archivos:
        print("❌ ERROR: No hay archivos 'impos'.")
        return

    final_dfs = []
    
    for file_name in archivos:
        try:
            print(f"Procesando {file_name}...")
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                wb = load_workbook(file_name, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data = list(sheet.iter_rows(values_only=True))
                    if len(data) < 2: continue
                    
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df.columns = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(df.columns)]
                    
                    mapping = {}
                    for col in df.columns:
                        c_up = str(col).upper()
                        if 'CLIENTE' in c_up and 'CLIENTE' not in mapping.values(): mapping[col] = 'CLIENTE'
                        if ('NUMERO' in c_up or 'CONTROL' in c_up) and 'CONTROL' not in mapping.values(): mapping[col] = 'CONTROL'
                        if 'ADUANA' in c_up and 'ADUANA' not in mapping.values(): mapping[col] = 'ADUANA'
                        if 'FECHA' in c_up and 'FECHA' not in mapping.values(): mapping[col] = 'FECHA'
                        if 'REFERENCIA' in c_up and 'REFERENCIA' not in mapping.values(): mapping[col] = 'REFERENCIA'
                        if 'ANTIDROGA' in c_up and 'ANTIDROGA' not in mapping.values(): mapping[col] = 'ANTIDROGA'
                        if 'CHOFER' in c_up and 'CHOFER' not in mapping.values(): mapping[col] = 'CHOFER'
                    
                    if sheet_name == '2023' and 'CLIENTE' not in mapping.values():
                         df = df.rename(columns={df.columns[0]:'CLIENTE', df.columns[1]:'CONTROL', df.columns[2]:'ADUANA', df.columns[3]:'FECHA'})
                    else:
                         df = df.rename(columns=mapping)

                    required = ['CLIENTE', 'CONTROL', 'ADUANA', 'FECHA']
                    if all(c in df.columns for c in required):
                        cols_to_keep = required + [c for c in ['REFERENCIA', 'ANTIDROGA', 'CHOFER'] if c in df.columns]
                        df = df[cols_to_keep]
                        df['CONTROL'] = df['CONTROL'].astype(str).str.replace(r'\.0$', '', regex=True)
                        final_dfs.append(df)
            
            elif file_name.endswith('.csv'):
                df = pd.read_csv(file_name, low_memory=False)
                df = df.rename(columns={'NUMERO de C': 'CONTROL', 'NUMERO DE C': 'CONTROL'})
                required = ['CLIENTE', 'CONTROL', 'ADUANA', 'FECHA']
                if all(c in df.columns for c in required):
                    df['CONTROL'] = df['CONTROL'].astype(str).str.replace(r'\.0$', '', regex=True)
                    final_dfs.append(df)
        
        except Exception as e:
            print(f"❌ Error procesando {file_name}: {e}")

    if not final_dfs:
        print("❌ ERROR: No hay datos.")
        return

    df_total = pd.concat([d.reset_index(drop=True) for d in final_dfs], ignore_index=True, sort=False)
    
    # Filtro
    df_total['ADUANA'] = df_total['ADUANA'].astype(str)
    df_total = df_total[df_total['ADUANA'].str.contains('TIENDITAS|S/ANTONIO|UREÑA', na=False, case=False)]

    # Robust Date Processing
    def robust_to_datetime(s):
        if pd.isna(s): return pd.NaT
        try:
            # First try default
            return pd.to_datetime(s, errors='raise')
        except:
            try:
                # Try dayfirst
                return pd.to_datetime(s, dayfirst=True, errors='raise')
            except:
                return pd.NaT

    df_total['FECHA_DT'] = df_total['FECHA'].apply(robust_to_datetime)
    
    meses = {1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio', 
             7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}
    df_total['MES'] = df_total['FECHA_DT'].dt.month.map(meses).fillna('Sin Fecha')

    df_total.to_json('datos_finales.json', orient='records', force_ascii=False)
    print(f"✅ ¡ÉXITO! {len(df_total)} registros consolidados.")

if __name__ == "__main__":
    limpiar_datos()